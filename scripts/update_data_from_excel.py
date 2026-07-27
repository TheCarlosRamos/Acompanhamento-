import argparse
import html
import json
import math
import re
import shutil
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Dependencia ausente: instale openpyxl com `pip install openpyxl`."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "Planilha Modelo setembro25.xlsx"
DEFAULT_COMPLEMENT = ROOT / "projects_full.xlsx"
PROJECTS_JSON = ROOT / "page" / "ppi_landing_site_v2" / "data" / "projects_full.json"
METRICS_JSON = ROOT / "page" / "ppi_landing_site_v2" / "data" / "metrics.json"
COMPLETE_JSON = ROOT / "projetos_completos.json"

PROJECT_COLUMNS = [
    "guid",
    "nome_projeto",
    "subsecretaria",
    "status_atual_do_projeto",
    "questoes_chaves",
    "proximas_etapas",
    "status_dos_estudos",
    "status_consulta_publica",
    "status_do_tcu",
    "status_do_edital",
    "status_do_leilao",
    "status_do_contrato",
    "descricao_do_projeto",
    "nome_completo",
    "descricao_curta",
    "setor",
    "subsetor",
    "organizacao",
    "localizacoes",
    "latitude",
    "longitude",
    "endereco_principal",
    "custo_estimado",
    "moeda",
    "custo_original",
    "status_atividade",
    "eh_ppp",
    "tipo_projeto",
    "arranjo_contratual",
    "processo_licitacao",
    "outro_arranjo_contratual",
    "outro_processo_licitacao",
    "uf",
    "permalink",
]

FIELD_ALIASES = {
    "subsecretaria": ["2000720", "subsecretaria", "secretaria_do_ppi_responsavel_pelo_projeto"],
    "status_atual_do_projeto": ["2000726", "status atual do projeto", "status_atual_do_projeto", "Status do projeto"],
    "questoes_chaves": ["2000727", "questoes chaves", "questões chaves", "questoes_chaves"],
    "proximas_etapas": ["2000728", "proximas etapas do projeto", "próximas etapas do projeto", "proximas_etapas"],
    "status_dos_estudos": ["2001218", "status dos estudos", "status_dos_estudos"],
    "status_consulta_publica": ["2001221", "status consulta publica", "status consulta pública", "status_consulta_publica"],
    "status_do_tcu": ["2001224", "status do tcu", "status_do_tcu"],
    "status_do_edital": ["2001226", "status do edital", "status_do_edital"],
    "status_do_leilao": ["2001229", "status do leilao", "status do leilão", "status_do_leilao"],
    "status_do_contrato": ["2001230", "status do contrato", "status_do_contrato"],
    "descricao_do_projeto": ["2001232", "descricao do projeto", "descrição do projeto", "descricao_do_projeto", "informacoes_do_projeto"],
}

BASE_ALIASES = {
    "id": ["ID", "id"],
    "nome_completo": ["Title", "titulo", "título", "nome_completo"],
    "descricao_curta": ["informacoes_do_projeto", "descricao_curta", "descrição curta", "descricao curta"],
    "setor": ["Setores", "setor"],
    "subsetor": ["Subsetores", "subsetor"],
    "organizacao": ["orgaos_envolvidos", "organizacao", "organização"],
    "uf": ["UF", "uf"],
    "permalink": ["Permalink", "permalink"],
    "status_atividade": ["projeto_ativo", "status_atividade"],
    "custo_estimado": ["vl_estimadosdivulgados_potenciais", "numero_capex", "custo_estimado"],
}

PHASE_FIELDS = {
    "estudo": "status_dos_estudos",
    "estudos": "status_dos_estudos",
    "consulta publica": "status_consulta_publica",
    "consulta publica": "status_consulta_publica",
    "acordao tcu": "status_do_tcu",
    "tcu": "status_do_tcu",
    "edital": "status_do_edital",
    "leilao de projeto": "status_do_leilao",
    "leilao": "status_do_leilao",
    "contrato": "status_do_contrato",
}


def normalize_key(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\s+", " ", text.replace("_", " ")).strip().lower()
    return text


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.replace("_x000D_", "\n")
        value = re.sub(r"[ \t]+", " ", value.replace("\r\n", "\n").replace("\r", "\n")).strip()
        return value or None
    return value


def html_to_text(value):
    value = clean_value(value)
    if not isinstance(value, str):
        return value
    if "<" not in value and "&" not in value:
        return value

    text = html.unescape(value)
    text = re.sub(r"(?is)<\s*script[^>]*>.*?<\s*/\s*script\s*>", " ", text)
    text = re.sub(r"(?is)<\s*style[^>]*>.*?<\s*/\s*style\s*>", " ", text)
    text = re.sub(r"(?is)<!--.*?-->", " ", text)
    text = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", text)
    text = re.sub(r"(?i)<\s*/\s*(p|div|h[1-6]|tr|table|ul|ol)\s*>", "\n\n", text)
    text = re.sub(r"(?i)<\s*(p|div|h[1-6]|tr|table|ul|ol)[^>]*>", "\n\n", text)
    text = re.sub(r"(?i)<\s*li[^>]*>", "\n- ", text)
    text = re.sub(r"(?i)<\s*/\s*li\s*>", "\n", text)
    text = re.sub(r"(?i)<\s*/?\s*(td|th)[^>]*>", " ", text)
    text = re.sub(r"(?is)<[^>]+>", "", text)
    text = html.unescape(text).replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip() or None


def clean_project_text(project):
    fields = [
        "descricao_do_projeto",
        "descricao_curta",
        "status_atual_do_projeto",
        "questoes_chaves",
        "proximas_etapas",
    ]
    for field in fields:
        project[field] = html_to_text(project.get(field))
    return project


def number_or_none(value):
    value = clean_value(value)
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def first_present(row, aliases, header_map):
    for alias in aliases:
        idx = header_map.get(normalize_key(alias))
        if idx is not None:
            value = clean_value(row[idx])
            if value not in (None, ""):
                return value
    return None


def row_dict_from_sheet(row, headers):
    return {
        str(header): clean_value(row[idx]) if idx < len(row) else None
        for idx, header in enumerate(headers)
        if header
    }


def find_sheet(workbook, preferred_name):
    if preferred_name and preferred_name in workbook.sheetnames:
        return workbook[preferred_name]
    if "Planilha" in workbook.sheetnames:
        return workbook["Planilha"]
    return workbook[workbook.sheetnames[0]]


def load_existing_projects(path):
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if isinstance(data, dict):
        projects = data.get("projetos", [])
    else:
        projects = data
    by_title = {}
    for project in projects:
        title = clean_value(project.get("nome_completo") or project.get("nome_projeto"))
        if title:
            by_title[normalize_key(title)] = project
    return by_title


def load_complement_projects(path):
    if not path or not path.exists():
        return {}, {"path": str(path) if path else None, "rows": 0, "matched": 0}

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Projetos"] if "Projetos" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_value(value) for value in next(rows)]
    header_map = {normalize_key(header): idx for idx, header in enumerate(headers) if header}

    aliases = {
        "guid": ["GUID", "guid"],
        "nome_projeto": ["Nome do Projeto", "nome_projeto"],
        "nome_completo": ["Nome Completo", "nome_completo"],
        "status_atual_do_projeto": ["Status Atual do Projeto", "status_atual_do_projeto"],
        "questoes_chaves": ["Questões Chaves", "questoes_chaves"],
        "proximas_etapas": ["Próximas Etapas", "proximas_etapas"],
        "latitude": ["Latitude", "latitude"],
        "longitude": ["Longitude", "longitude"],
    }

    by_title = {}
    row_count = 0
    for row in rows:
        if not any(clean_value(value) not in (None, "") for value in row):
            continue
        row_count += 1
        item = {}
        for field, field_aliases in aliases.items():
            item[field] = first_present(row, field_aliases, header_map)
        item["status_atual_do_projeto"] = html_to_text(item.get("status_atual_do_projeto"))
        item["questoes_chaves"] = html_to_text(item.get("questoes_chaves"))
        item["proximas_etapas"] = html_to_text(item.get("proximas_etapas"))
        item["latitude"] = number_or_none(item.get("latitude"))
        item["longitude"] = number_or_none(item.get("longitude"))

        for title in [item.get("nome_completo"), item.get("nome_projeto")]:
            key = normalize_key(title)
            if key:
                by_title[key] = item

    return by_title, {"path": str(path), "rows": row_count, "matched": 0}


def apply_complement(projects, complement_by_title, complement_stats):
    if not complement_by_title:
        return complement_stats

    matched = 0
    updated = Counter()
    for project in projects:
        keys = [
            normalize_key(project.get("nome_completo")),
            normalize_key(project.get("nome_projeto")),
        ]
        complement = next((complement_by_title[key] for key in keys if key in complement_by_title), None)
        if not complement:
            continue
        matched += 1

        for field in ["status_atual_do_projeto", "questoes_chaves", "proximas_etapas"]:
            value = complement.get(field)
            if value not in (None, ""):
                project[field] = value
                updated[field] += 1

        lat = complement.get("latitude")
        lon = complement.get("longitude")
        if lat is not None and lon is not None:
            project["latitude"] = lat
            project["longitude"] = lon
            updated["coordenadas"] += 1

        clean_project_text(project)

    complement_stats["matched"] = matched
    complement_stats["updated"] = dict(updated)
    return complement_stats


def derive_phase_status(row_values):
    status_by_field = {}
    for idx in range(6):
        andamento = normalize_key(row_values.get(f"etapas_{idx}_andamento"))
        status = clean_value(row_values.get(f"etapas_{idx}_status"))
        if not andamento or not status:
            continue
        for label, target in PHASE_FIELDS.items():
            if label in andamento:
                status_by_field[target] = status
                break
    return status_by_field


def first_coordinates(row_values):
    for idx in range(34):
        lat = number_or_none(row_values.get(f"pontos_{idx}_latitude"))
        lon = number_or_none(row_values.get(f"pontos_{idx}_longitude"))
        if lat is not None and lon is not None:
            return lat, lon
    return None, None


def build_projects(excel_path, sheet_name=None, complement_path=None):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = find_sheet(workbook, sheet_name)
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_value(value) for value in next(rows)]
    header_map = {normalize_key(header): idx for idx, header in enumerate(headers) if header}
    existing_by_title = load_existing_projects(PROJECTS_JSON)

    projects = []
    for row in rows:
        if not any(clean_value(value) not in (None, "") for value in row):
            continue
        row_values = row_dict_from_sheet(row, headers)

        title = first_present(row, BASE_ALIASES["nome_completo"], header_map)
        project_id = first_present(row, BASE_ALIASES["id"], header_map)
        if not title and not project_id:
            continue

        existing = existing_by_title.get(normalize_key(title))
        guid = (existing or {}).get("guid") or (f"excel-{project_id}" if project_id is not None else normalize_key(title))
        lat, lon = first_coordinates(row_values)
        phase_status = derive_phase_status(row_values)

        project = {column: None for column in PROJECT_COLUMNS}
        project.update(
            {
                "guid": str(guid),
                "nome_projeto": title or f"Projeto {project_id}",
                "nome_completo": title or f"Projeto {project_id}",
                "descricao_curta": first_present(row, BASE_ALIASES["descricao_curta"], header_map),
                "setor": first_present(row, BASE_ALIASES["setor"], header_map),
                "subsetor": first_present(row, BASE_ALIASES["subsetor"], header_map),
                "organizacao": first_present(row, BASE_ALIASES["organizacao"], header_map),
                "localizacoes": first_present(row, BASE_ALIASES["uf"], header_map),
                "latitude": lat,
                "longitude": lon,
                "endereco_principal": first_present(row, BASE_ALIASES["uf"], header_map),
                "custo_estimado": number_or_none(first_present(row, BASE_ALIASES["custo_estimado"], header_map)),
                "moeda": "BRL" if number_or_none(first_present(row, BASE_ALIASES["custo_estimado"], header_map)) is not None else None,
                "status_atividade": first_present(row, BASE_ALIASES["status_atividade"], header_map),
                "uf": first_present(row, BASE_ALIASES["uf"], header_map),
                "permalink": first_present(row, BASE_ALIASES["permalink"], header_map),
            }
        )

        for field, aliases in FIELD_ALIASES.items():
            project[field] = first_present(row, aliases, header_map) or phase_status.get(field)

        if not project["descricao_do_projeto"]:
            project["descricao_do_projeto"] = project["descricao_curta"]

        clean_project_text(project)
        projects.append(project)

    complement_stats = {"path": None, "rows": 0, "matched": 0}
    if complement_path:
        complement_by_title, complement_stats = load_complement_projects(complement_path)
        complement_stats = apply_complement(projects, complement_by_title, complement_stats)

    projects.sort(key=lambda item: normalize_key(item.get("nome_completo") or item.get("guid")))
    return projects, sheet.title, complement_stats


def last_completed_phase(project):
    phases = [
        ("Estudos", "status_dos_estudos"),
        ("Consulta Pública", "status_consulta_publica"),
        ("TCU", "status_do_tcu"),
        ("Edital", "status_do_edital"),
        ("Leilão", "status_do_leilao"),
        ("Contrato", "status_do_contrato"),
    ]
    last = "Nenhuma"
    for label, field in phases:
        value = normalize_key(project.get(field))
        if any(marker in value for marker in ["concluido", "completed", "assinado", "assinatura"]):
            last = label
    return last


def stats(values):
    values = sorted(values)
    if not values:
        return None
    middle = len(values) // 2
    if len(values) % 2:
        median = values[middle]
    else:
        median = (values[middle - 1] + values[middle]) / 2
    return {
        "projetos_com_custo": len(values),
        "soma": sum(values),
        "min": min(values),
        "max": max(values),
        "media": sum(values) / len(values),
        "mediana": median,
    }


def build_metrics(projects, source_name):
    costs = [project["custo_estimado"] for project in projects if isinstance(project.get("custo_estimado"), (int, float))]
    costs_by_sector = Counter()
    for project in projects:
        value = project.get("custo_estimado")
        if isinstance(value, (int, float)):
            costs_by_sector[project.get("setor") or "Sem setor"] += value

    top10 = sorted(
        [
            {
                "guid": project.get("guid"),
                "nome": project.get("nome_completo") or project.get("nome_projeto"),
                "setor": project.get("setor"),
                "valor": project.get("custo_estimado"),
            }
            for project in projects
            if isinstance(project.get("custo_estimado"), (int, float))
        ],
        key=lambda item: item["valor"],
        reverse=True,
    )[:10]

    return {
        "fonte": source_name,
        "total_projetos": len(projects),
        "projetos_com_coordenadas": sum(1 for project in projects if project.get("latitude") is not None and project.get("longitude") is not None),
        "por_setor": dict(Counter(project.get("setor") or "Sem setor" for project in projects)),
        "por_ultima_etapa_concluida": dict(Counter(last_completed_phase(project) for project in projects)),
        "por_situacao_atual": dict(Counter(project.get("status_atual_do_projeto") or "Nao informado" for project in projects)),
        "custos_por_moeda": {"BRL": stats(costs)} if costs else {},
        "custos_por_setor_por_moeda": {"BRL": dict(costs_by_sector)} if costs_by_sector else {},
        "por_uf": dict(Counter(project.get("uf") or project.get("localizacoes") or "Nao informado" for project in projects)),
        "top10_custo_brl": top10,
        "data_geracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def backup(path):
    if not path.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = path.with_suffix(path.suffix + f".bak-{stamp}")
    shutil.copy2(path, backup_path)
    return backup_path


def main():
    parser = argparse.ArgumentParser(description="Atualiza os JSONs da aplicacao a partir da planilha Excel da raiz.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Caminho da planilha .xlsx.")
    parser.add_argument("--sheet", default=None, help="Nome da aba. Padrao: Planilha, se existir.")
    parser.add_argument("--complement", default=str(DEFAULT_COMPLEMENT), help="Planilha complementar para situacao, passos, pontos e coordenadas.")
    parser.add_argument("--no-complement", action="store_true", help="Nao usar a planilha complementar.")
    parser.add_argument("--no-backup", action="store_true", help="Nao criar backup dos JSONs antes de atualizar.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = ROOT / excel_path
    if not excel_path.exists():
        raise SystemExit(f"Planilha nao encontrada: {excel_path}")

    complement_path = None
    if not args.no_complement:
        complement_path = Path(args.complement)
        if not complement_path.is_absolute():
            complement_path = ROOT / complement_path

    projects, sheet_name, complement_stats = build_projects(excel_path, args.sheet, complement_path)
    if not projects:
        raise SystemExit("Nenhum projeto foi encontrado na planilha.")

    if not args.no_backup:
        for path in [PROJECTS_JSON, METRICS_JSON, COMPLETE_JSON]:
            backup(path)

    complete = {
        "metadados": {
            "total_projetos": len(projects),
            "total_colunas": len(PROJECT_COLUMNS),
            "data_geracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fonte": excel_path.name,
            "aba": sheet_name,
            "complemento": complement_stats,
            "colunas": PROJECT_COLUMNS,
        },
        "projetos": projects,
    }

    write_json(PROJECTS_JSON, projects)
    write_json(METRICS_JSON, build_metrics(projects, excel_path.name))
    write_json(COMPLETE_JSON, complete)

    print(f"Planilha: {excel_path}")
    print(f"Aba: {sheet_name}")
    if complement_stats.get("path"):
        print(f"Complemento: {complement_stats['path']}")
        print(f"Linhas do complemento: {complement_stats.get('rows', 0)}")
        print(f"Projetos casados com complemento: {complement_stats.get('matched', 0)}")
        print(f"Campos atualizados pelo complemento: {complement_stats.get('updated', {})}")
    print(f"Projetos processados: {len(projects)}")
    print(f"Atualizado: {PROJECTS_JSON}")
    print(f"Atualizado: {METRICS_JSON}")
    print(f"Atualizado: {COMPLETE_JSON}")


if __name__ == "__main__":
    main()
